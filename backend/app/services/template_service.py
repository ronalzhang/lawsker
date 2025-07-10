"""
标准模板服务
提供各种业务数据的标准Excel模板生成和下载功能
"""

import os
import io
from typing import Dict, Any, List, Optional
from datetime import datetime
from fastapi import HTTPException
from fastapi.responses import StreamingResponse
import logging

# 尝试导入pandas和openpyxl
try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

logger = logging.getLogger(__name__)

class TemplateService:
    """标准模板服务"""
    
    def __init__(self):
        """初始化模板服务"""
        # 债务催收标准模板字段
        self.debt_collection_template = {
            'sheet_name': '债务催收数据模板',
            'required_fields': [
                '债务人姓名', '身份证号', '债务人电话', '欠款金额', '逾期天数'
            ],
            'optional_fields': [
                '亲属电话', '紧急联系人电话', '其他联系人电话', '银行名称', '贷款类型',
                '合同编号', '最后还款日期', '债务人地址', '工作单位', '月收入',
                '担保人姓名', '担保人电话', '备注'
            ],
            'sample_data': [
                {
                    '债务人姓名': '张三',
                    '身份证号': '110101199001011234',
                    '债务人电话': '13800138000',
                    '亲属电话': '13900139000,13700137000',
                    '紧急联系人电话': '13600136000',
                    '其他联系人电话': '13500135000',
                    '欠款金额': '50000',
                    '逾期天数': '60',
                    '银行名称': '某某银行',
                    '贷款类型': '个人消费贷',
                    '合同编号': 'HT202401001',
                    '最后还款日期': '2023-11-15',
                    '债务人地址': '北京市朝阳区某某街道123号',
                    '工作单位': '某科技有限公司',
                    '月收入': '8000',
                    '担保人姓名': '李四',
                    '担保人电话': '13400134000',
                    '备注': '优质客户，有还款意愿'
                },
                {
                    '债务人姓名': '王五',
                    '身份证号': '220101199502022345',
                    '债务人电话': '13801138001',
                    '亲属电话': '13901139001',
                    '紧急联系人电话': '',
                    '其他联系人电话': '',
                    '欠款金额': '35000',
                    '逾期天数': '45',
                    '银行名称': '某某银行',
                    '贷款类型': '小额信贷',
                    '合同编号': 'HT202401002',
                    '最后还款日期': '2023-12-01',
                    '债务人地址': '上海市浦东新区某某路456号',
                    '工作单位': '某贸易公司',
                    '月收入': '6500',
                    '担保人姓名': '',
                    '担保人电话': '',
                    '备注': '联系困难，需多次催收'
                },
                {
                    '债务人姓名': '赵六',
                    '身份证号': '330101199203033456',
                    '债务人电话': '13802138002',
                    '亲属电话': '13902139002,13702137002',
                    '紧急联系人电话': '13602136002',
                    '其他联系人电话': '',
                    '欠款金额': '120000',
                    '逾期天数': '90',
                    '银行名称': '某某银行',
                    '贷款类型': '个人住房贷款',
                    '合同编号': 'HT202401003',
                    '最后还款日期': '2023-10-20',
                    '债务人地址': '杭州市西湖区某某小区789号',
                    '工作单位': '某建筑公司',
                    '月收入': '12000',
                    '担保人姓名': '孙七',
                    '担保人电话': '13403134003',
                    '备注': '房贷逾期，有房产可执行'
                }
            ],
            'field_descriptions': {
                '债务人姓名': '债务人的真实姓名（必填）',
                '身份证号': '18位身份证号码，格式：110101199001011234（必填）',
                '债务人电话': '债务人手机号，11位数字，格式：13800138000（必填）',
                '亲属电话': '债务人亲属电话，多个号码用英文逗号分隔（可选）',
                '紧急联系人电话': '紧急联系人电话，多个号码用英文逗号分隔（可选）',
                '其他联系人电话': '其他联系人电话，多个号码用英文逗号分隔（可选）',
                '欠款金额': '欠款金额，纯数字，单位：元（必填）',
                '逾期天数': '逾期天数，整数，单位：天（必填）',
                '银行名称': '放贷机构名称（可选）',
                '贷款类型': '贷款产品类型，如：个人消费贷、住房贷款等（可选）',
                '合同编号': '借款合同编号（可选）',
                '最后还款日期': '最后一次还款日期，格式：2023-11-15（可选）',
                '债务人地址': '债务人详细联系地址（可选）',
                '工作单位': '债务人工作单位名称（可选）',
                '月收入': '债务人月收入，数字，单位：元（可选）',
                '担保人姓名': '担保人姓名（可选）',
                '担保人电话': '担保人联系电话（可选）',
                '备注': '其他补充信息和特殊说明（可选）'
            }
        }
        
        # 其他业务模板可以在这里添加
        self.templates = {
            'debt_collection': self.debt_collection_template
        }
    
    async def generate_template(self, template_type: str = 'debt_collection') -> io.BytesIO:
        """
        生成标准Excel模板
        
        Args:
            template_type: 模板类型，默认为债务催收
            
        Returns:
            Excel文件的BytesIO对象
        """
        if not HAS_PANDAS:
            raise HTTPException(status_code=500, detail="系统不支持Excel文件生成")
        
        if template_type not in self.templates:
            raise HTTPException(status_code=400, detail=f"不支持的模板类型: {template_type}")
        
        template_config = self.templates[template_type]
        
        try:
            # 创建Excel工作簿
            wb = openpyxl.Workbook()
            
            # 删除默认工作表
            if wb.active:
                wb.remove(wb.active)
            
            # 创建主数据表
            main_sheet = wb.create_sheet(title=template_config['sheet_name'])
            
            # 创建说明表
            instruction_sheet = wb.create_sheet(title='填写说明')
            
            # 填充主数据表
            await self._fill_main_sheet(main_sheet, template_config)
            
            # 填充说明表
            await self._fill_instruction_sheet(instruction_sheet, template_config)
            
            # 保存到BytesIO
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            logger.info(f"成功生成{template_type}模板")
            return excel_buffer
            
        except Exception as e:
            logger.error(f"生成模板失败: {str(e)}")
            raise HTTPException(status_code=500, detail=f"生成模板失败: {str(e)}")
    
    async def _fill_main_sheet(self, sheet, template_config: Dict[str, Any]):
        """填充主数据表"""
        # 所有字段（必填 + 可选）
        all_fields = template_config['required_fields'] + template_config['optional_fields']
        
        # 定义样式
        header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        header_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        required_font = Font(name='微软雅黑', size=12, bold=True, color='FF0000')
        required_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
        
        data_font = Font(name='微软雅黑', size=11)
        data_alignment = Alignment(horizontal='left', vertical='center')
        data_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入表头
        for col_idx, field in enumerate(all_fields, 1):
            cell = sheet.cell(row=1, column=col_idx)
            
            # 必填字段标记
            if field in template_config['required_fields']:
                cell.value = f"{field} *"
                cell.font = required_font
                cell.fill = required_fill
            else:
                cell.value = field
                cell.font = header_font
                cell.fill = header_fill
            
            cell.alignment = header_alignment
            cell.border = header_border
            
            # 设置列宽
            if '电话' in field:
                sheet.column_dimensions[get_column_letter(col_idx)].width = 15
            elif '地址' in field or '备注' in field:
                sheet.column_dimensions[get_column_letter(col_idx)].width = 25
            elif '姓名' in field:
                sheet.column_dimensions[get_column_letter(col_idx)].width = 12
            elif '金额' in field or '收入' in field:
                sheet.column_dimensions[get_column_letter(col_idx)].width = 12
            else:
                sheet.column_dimensions[get_column_letter(col_idx)].width = 18
        
        # 写入示例数据
        for row_idx, sample_row in enumerate(template_config['sample_data'], 2):
            for col_idx, field in enumerate(all_fields, 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = sample_row.get(field, '')
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = data_border
        
        # 冻结首行
        sheet.freeze_panes = 'A2'
        
        # 添加数据验证（对关键字段）
        await self._add_data_validation(sheet, all_fields, len(template_config['sample_data']) + 10)
    
    async def _add_data_validation(self, sheet, fields: List[str], max_row: int):
        """添加数据验证"""
        try:
            from openpyxl.worksheet.datavalidation import DataValidation
            
            # 身份证号验证
            if '身份证号' in fields:
                col_idx = fields.index('身份证号') + 1
                col_letter = get_column_letter(col_idx)
                
                id_validation = DataValidation(
                    type="custom",
                    formula1=f'AND(LEN({col_letter}2)=18,ISNUMBER(VALUE(LEFT({col_letter}2,17))))',
                    errorTitle="身份证号格式错误",
                    error="请输入18位身份证号码"
                )
                sheet.add_data_validation(id_validation)
                id_validation.add(f'{col_letter}2:{col_letter}{max_row}')
            
            # 手机号验证
            phone_fields = [field for field in fields if '电话' in field]
            for phone_field in phone_fields:
                if phone_field in fields:
                    col_idx = fields.index(phone_field) + 1
                    col_letter = get_column_letter(col_idx)
                    
                    phone_validation = DataValidation(
                        type="custom", 
                        formula1=f'OR(LEN({col_letter}2)=0,AND(LEN({col_letter}2)=11,LEFT({col_letter}2,1)="1"))',
                        errorTitle="手机号格式错误",
                        error="请输入11位手机号码，以1开头"
                    )
                    sheet.add_data_validation(phone_validation)
                    phone_validation.add(f'{col_letter}2:{col_letter}{max_row}')
            
            # 金额验证
            if '欠款金额' in fields:
                col_idx = fields.index('欠款金额') + 1
                col_letter = get_column_letter(col_idx)
                
                amount_validation = DataValidation(
                    type="decimal",
                    operator="greaterThan",
                    formula1=0,
                    errorTitle="金额格式错误",
                    error="金额必须大于0"
                )
                sheet.add_data_validation(amount_validation)
                amount_validation.add(f'{col_letter}2:{col_letter}{max_row}')
            
            # 逾期天数验证
            if '逾期天数' in fields:
                col_idx = fields.index('逾期天数') + 1
                col_letter = get_column_letter(col_idx)
                
                days_validation = DataValidation(
                    type="whole",
                    operator="between",
                    formula1=0,
                    formula2=3650,
                    errorTitle="逾期天数错误",
                    error="逾期天数必须在0-3650之间"
                )
                sheet.add_data_validation(days_validation)
                days_validation.add(f'{col_letter}2:{col_letter}{max_row}')
        
        except Exception as e:
            logger.warning(f"添加数据验证失败: {str(e)}")
    
    async def _fill_instruction_sheet(self, sheet, template_config: Dict[str, Any]):
        """填充说明表"""
        # 标题样式
        title_font = Font(name='微软雅黑', size=16, bold=True, color='2F5597')
        subtitle_font = Font(name='微软雅黑', size=14, bold=True, color='4F81BD')
        text_font = Font(name='微软雅黑', size=11)
        required_font = Font(name='微软雅黑', size=11, color='FF0000')
        
        # 写入标题
        sheet['A1'] = '债务催收数据模板填写说明'
        sheet['A1'].font = title_font
        sheet.merge_cells('A1:D1')
        
        # 重要提示
        sheet['A3'] = '⚠️ 重要提示：'
        sheet['A3'].font = subtitle_font
        
        warnings = [
            '1. 标有 * 号的字段为必填项，不能为空',
            '2. 身份证号必须为18位，格式正确',
            '3. 手机号必须为11位，以1开头',
            '4. 多个电话号码用英文逗号(,)分隔',
            '5. 金额字段请填入纯数字，不要包含货币符号',
            '6. 请确保数据真实有效，以提高催收成功率'
        ]
        
        for i, warning in enumerate(warnings, 4):
            sheet[f'A{i}'] = warning
            sheet[f'A{i}'].font = required_font
        
        # 字段说明
        sheet[f'A{len(warnings) + 5}'] = '📋 字段说明：'
        sheet[f'A{len(warnings) + 5}'].font = subtitle_font
        
        # 表头
        start_row = len(warnings) + 6
        sheet[f'A{start_row}'] = '字段名称'
        sheet[f'B{start_row}'] = '是否必填'
        sheet[f'C{start_row}'] = '说明'
        
        for col in ['A', 'B', 'C']:
            sheet[f'{col}{start_row}'].font = subtitle_font
        
        # 字段详细说明
        row_idx = start_row + 1
        for field in template_config['required_fields'] + template_config['optional_fields']:
            sheet[f'A{row_idx}'] = field
            sheet[f'B{row_idx}'] = '必填' if field in template_config['required_fields'] else '可选'
            sheet[f'C{row_idx}'] = template_config['field_descriptions'].get(field, '')
            
            # 设置字体
            sheet[f'A{row_idx}'].font = text_font
            if field in template_config['required_fields']:
                sheet[f'B{row_idx}'].font = required_font
            else:
                sheet[f'B{row_idx}'].font = text_font
            sheet[f'C{row_idx}'].font = text_font
            
            row_idx += 1
        
        # 设置列宽
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 50
        
        # 联系信息
        contact_row = row_idx + 2
        sheet[f'A{contact_row}'] = '📞 如有疑问，请联系技术支持'
        sheet[f'A{contact_row}'].font = subtitle_font
        
        sheet[f'A{contact_row + 1}'] = '邮箱：support@lawsker.com'
        sheet[f'A{contact_row + 1}'].font = text_font
        
        sheet[f'A{contact_row + 2}'] = '电话：400-123-4567'
        sheet[f'A{contact_row + 2}'].font = text_font
    
    async def download_template(self, template_type: str = 'debt_collection') -> StreamingResponse:
        """
        下载标准模板
        
        Args:
            template_type: 模板类型
            
        Returns:
            StreamingResponse对象用于文件下载
        """
        try:
            # 生成模板
            excel_buffer = await self.generate_template(template_type)
            
            # 生成文件名
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            if template_type == 'debt_collection':
                filename = f'债务催收数据模板_{timestamp}.xlsx'
            else:
                filename = f'{template_type}_模板_{timestamp}.xlsx'
            
            # 返回下载响应
            return StreamingResponse(
                io.BytesIO(excel_buffer.read()),
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={
                    'Content-Disposition': f'attachment; filename="{filename}"'
                }
            )
            
        except Exception as e:
            logger.error(f"下载模板失败: {str(e)}")
            raise HTTPException(status_code=500, detail=f"下载模板失败: {str(e)}")
    
    def get_template_info(self, template_type: str = 'debt_collection') -> Dict[str, Any]:
        """获取模板信息"""
        if template_type not in self.templates:
            raise HTTPException(status_code=400, detail=f"不支持的模板类型: {template_type}")
        
        template_config = self.templates[template_type]
        
        return {
            'template_type': template_type,
            'template_name': template_config['sheet_name'],
            'required_fields': template_config['required_fields'],
            'optional_fields': template_config['optional_fields'],
            'total_fields': len(template_config['required_fields']) + len(template_config['optional_fields']),
            'sample_records': len(template_config['sample_data']),
            'field_descriptions': template_config['field_descriptions'],
            'download_url': f'/api/v1/upload/template/download?type={template_type}'
        }
    
    def get_available_templates(self) -> List[Dict[str, Any]]:
        """获取所有可用模板"""
        templates = []
        
        for template_type, config in self.templates.items():
            templates.append({
                'type': template_type,
                'name': config['sheet_name'],
                'description': f"包含{len(config['required_fields'])}个必填字段和{len(config['optional_fields'])}个可选字段",
                'required_fields_count': len(config['required_fields']),
                'optional_fields_count': len(config['optional_fields']),
                'download_url': f'/api/v1/upload/template/download?type={template_type}'
            })
        
        return templates 